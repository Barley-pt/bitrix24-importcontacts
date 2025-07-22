import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import requests
import os

# Ask user whether to check for duplicates
def ask_duplicate_check():
    response = messagebox.askyesno(
        "Duplicate Check",
        "Do you want to check for duplicates before importing?\n\n(Recommended to avoid duplicate contacts)"
    )
    if not response:
        messagebox.showwarning(
            "Warning",
            "Duplicate check is disabled.\nAll contacts will be imported as new."
        )
    return response

# Fetch contact fields from Bitrix24
def fetch_bitrix_fields(webhook_url):
    response = requests.get(f"{webhook_url.rstrip('/')}/crm.contact.fields.json")
    result = response.json()
    if not result.get('result'):
        messagebox.showerror("Error", "Failed to fetch fields from Bitrix24.")
        return []

    fields = result['result']
    allowed_types = ['string', 'integer', 'double', 'boolean', 'enumeration', 'date', 'datetime']
    bitrix_fields = [key for key, val in fields.items() if val.get('type') in allowed_types and not val.get('isReadOnly', False)]
    field_labels = {key: val.get('title', key) for key in bitrix_fields}
    return bitrix_fields, field_labels

# 1. GUI to select Excel file
def select_file(check_duplicates):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        map_fields(file_path, headers, check_duplicates)

# 2. GUI to map Excel headers to Bitrix fields
def map_fields(file_path, headers, check_duplicates):
    webhook = simple_input("Enter your Bitrix24 Webhook URL:")
    field_keys, field_labels = fetch_bitrix_fields(webhook)

    def submit_mappings():
        mappings = {}
        for i, header in enumerate(headers):
            field = combo_vars[i].get()
            if field:
                mappings[header] = field
        window.destroy()
        run_import(file_path, mappings, webhook, check_duplicates)

    window = tk.Tk()
    window.title("Field Mapping")
    tk.Label(window, text="Map Excel Columns to Bitrix24 Contact Fields").grid(row=0, column=0, columnspan=2)

    combo_vars = []
    for i, header in enumerate(headers):
        tk.Label(window, text=header).grid(row=i+1, column=0)
        var = tk.StringVar(window)
        dropdown = tk.OptionMenu(window, var, *[""] + [f"{f} - {field_labels[f]}" for f in field_keys])
        dropdown.grid(row=i+1, column=1)
        combo_vars.append(var)

    tk.Button(window, text="Start Import", command=submit_mappings).grid(row=len(headers)+1, column=0, columnspan=2)
    window.mainloop()

# Simple text input dialog
def simple_input(prompt_text):
    def on_submit():
        nonlocal user_input
        user_input = entry.get()
        input_window.destroy()

    user_input = None
    input_window = tk.Tk()
    input_window.title("Input Required")
    tk.Label(input_window, text=prompt_text).pack()
    entry = tk.Entry(input_window, width=50)
    entry.pack()
    tk.Button(input_window, text="Submit", command=on_submit).pack()
    input_window.mainloop()
    return user_input

# Find contact by email or phone
def find_existing_contact(webhook, email=None, phone=None):
    filters = {}
    if email:
        filters["EMAIL"] = email
    if phone:
        filters["PHONE"] = phone
    if not filters:
        return None

    response = requests.post(
        f"{webhook.rstrip('/')}/crm.contact.list.json",
        json={"filter": filters, "select": ["ID"]}
    )
    result = response.json().get("result", [])
    return result[0]["ID"] if result else None

# Main import function
def run_import(file_path, mappings, webhook, check_duplicates):
    if not webhook.endswith("/"):
        webhook += "/"

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    bitrix_id_col = len(headers)
    sheet.cell(row=1, column=bitrix_id_col + 1).value = "BITRIX_ID"

    success_count = 0
    fail_count = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        contact_data = {"fields": {}}
        email = None
        phone = None

        for excel_col, bitrix_field_full in mappings.items():
            bitrix_field = bitrix_field_full.split(" - ")[0]
            value = row[headers.index(excel_col)]
            if value:
                if bitrix_field == "EMAIL":
                    email = value
                    contact_data["fields"][bitrix_field] = [{"VALUE": email, "VALUE_TYPE": "WORK"}]
                elif bitrix_field == "PHONE":
                    phone = value
                    contact_data["fields"][bitrix_field] = [{"VALUE": phone, "VALUE_TYPE": "WORK"}]
                else:
                    contact_data["fields"][bitrix_field] = value

        try:
            if check_duplicates:
                existing_id = find_existing_contact(webhook, email=email, phone=phone)
            else:
                existing_id = None

            if existing_id:
                print(f"Duplicate found. Using existing Contact ID: {existing_id}")
                sheet.cell(row=i, column=bitrix_id_col + 1).value = existing_id
            else:
                r = requests.post(f"{webhook}crm.contact.add.json", json=contact_data)
                r_json = r.json()
                new_id = r_json.get("result")
                if new_id:
                    print(f"Created new contact: {new_id}")
                    success_count += 1
                    sheet.cell(row=i, column=bitrix_id_col + 1).value = new_id
                else:
                    fail_count += 1
        except Exception as e:
            print(f"Error on row {i}: {e}")
            fail_count += 1

    dir_name, base_name = os.path.split(file_path)
    name_only, ext = os.path.splitext(base_name)
    new_file = os.path.join(dir_name, f"{name_only}_bitrix_imported{ext}")
    wb.save(new_file)

    messagebox.showinfo("Import Complete", f"✅ Success: {success_count}\n❌ Failed: {fail_count}\n💾 Saved: {new_file}")

# --- Main Program ---

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    check_duplicates = ask_duplicate_check()
    select_file(check_duplicates)
